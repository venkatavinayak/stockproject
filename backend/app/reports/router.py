from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime, timedelta
import io
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.app.database.session import get_db
from backend.app.reports.generator import build_pdf_report
from backend.app.models.transaction import Transaction
from backend.app.models.transaction_item import TransactionItem
from backend.app.models.expense import Expense
from backend.app.models.return import Return
from backend.app.models.product import Product
from backend.app.models.settings import StoreSettings
from backend.app.auth.deps import get_current_user
from backend.app.models.user import User

router = APIRouter(prefix="/reports", tags=["Reports & Exports"])

@router.get("/generate")
def generate_report(
    report_type: str,  # Daily, Weekly, Monthly, Yearly, Custom
    start_date: date,
    end_date: date,
    file_format: str = "pdf",  # pdf, excel, csv
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="End date cannot be prior to start date"
        )
        
    settings = db.query(StoreSettings).first() or StoreSettings(store_name="SmartStock AI")
    currency = settings.currency_symbol
    
    # 1. GENERATE PDF REPORT
    if file_format.lower() == "pdf":
        try:
            pdf_rel_path = build_pdf_report(db, report_type, start_date, end_date)
            absolute_pdf_path = os.path.abspath(pdf_rel_path)
            
            # Record notification of report generation
            notif = Notification(
                type="System",
                message=f"Business report ({report_type}) compiled in PDF format",
                timestamp=datetime.now()
            )
            db.add(notif)
            db.commit()
            
            return FileResponse(
                path=absolute_pdf_path,
                filename=os.path.basename(pdf_rel_path),
                media_type='application/pdf'
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")
            
    # 2. GENERATE EXCEL WORKBOOK (OpenPyXL)
    elif file_format.lower() == "excel":
        try:
            wb = openpyxl.Workbook()
            
            # Fonts and styles
            font_title = Font(name='Segoe UI', size=16, bold=True, color='1E3A8A')
            font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            font_bold = Font(name='Segoe UI', size=11, bold=True)
            font_regular = Font(name='Segoe UI', size=11)
            
            fill_header = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            fill_light = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            
            thin_side = Side(border_style="thin", color="D1D5DB")
            thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # --- Sheet 1: Executive Summary ---
            ws_summary = wb.active
            ws_summary.title = "Executive Summary"
            ws_summary.views.sheetView[0].showGridLines = True
            
            ws_summary["A1"] = f"{settings.store_name} - business performance"
            ws_summary["A1"].font = font_title
            
            ws_summary["A2"] = f"Period: {start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')}"
            ws_summary["A2"].font = font_regular
            
            # Fetch aggregates
            sales_data = db.query(
                func.sum(Transaction.grand_total).label("revenue"),
                func.sum(Transaction.buying_cost).label("cost"),
                func.sum(Transaction.profit).label("profit"),
                func.count(Transaction.id).label("bills")
            ).filter(Transaction.timestamp >= start_date, Transaction.timestamp <= datetime.combine(end_date, datetime.max.time())).first()
            
            rev = float(sales_data.revenue or 0.0)
            cost = float(sales_data.cost or 0.0)
            gross_profit = float(sales_data.profit or 0.0)
            bills = int(sales_data.bills or 0)
            
            expenses_val = db.query(func.sum(Expense.amount)).filter(Expense.date >= start_date, Expense.date <= end_date).scalar() or 0.0
            net_profit = gross_profit - expenses_val
            
            items_sold = db.query(func.sum(TransactionItem.quantity)).join(Transaction).filter(
                Transaction.timestamp >= start_date, Transaction.timestamp <= datetime.combine(end_date, datetime.max.time())
            ).scalar() or 0
            
            summary_metrics = [
                ("Total Revenue", rev, f"Total sales ({currency})"),
                ("Cost of Goods Sold (COGS)", cost, f"Supplier acquisition cost ({currency})"),
                ("Gross Profit", gross_profit, f"Revenue minus COGS ({currency})"),
                ("Logged Expenses", expenses_val, f"Rent, utility, internet bills ({currency})"),
                ("Net Profit", net_profit, f"Operational earnings ({currency})"),
                ("Total Bills", bills, "Count of sales transactions"),
                ("Total Items Sold", items_sold, "Count of items scanned")
            ]
            
            headers = ["Business Metric", "Value", "Notes"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=4, column=col_idx, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="left")
                
            for row_idx, (metric, val, note) in enumerate(summary_metrics, 5):
                c1 = ws_summary.cell(row=row_idx, column=1, value=metric)
                c2 = ws_summary.cell(row=row_idx, column=2, value=val)
                c3 = ws_summary.cell(row=row_idx, column=3, value=note)
                
                c1.font = font_regular
                c2.font = font_bold if "Profit" in metric or "Revenue" in metric else font_regular
                c3.font = font_regular
                
                c1.border = thin_border
                c2.border = thin_border
                c3.border = thin_border
                
                # Format numeric currency column
                if isinstance(val, float):
                    c2.number_format = f'"{currency}"#,##0.00'
                else:
                    c2.number_format = '#,##0'
                    
            # --- Sheet 2: Transactions ---
            ws_txs = wb.create_sheet(title="Transactions Ledger")
            ws_txs.views.sheetView[0].showGridLines = True
            
            tx_headers = ["Invoice No", "Timestamp", "Payment Method", "Items Count", "Subtotal", "GST Amount", "Discount", "Grand Total", "Profit"]
            for col_idx, header in enumerate(tx_headers, 1):
                cell = ws_txs.cell(row=1, column=col_idx, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center")
                
            txs = db.query(Transaction).filter(
                Transaction.timestamp >= start_date, Transaction.timestamp <= datetime.combine(end_date, datetime.max.time())
            ).all()
            
            for row_idx, t in enumerate(txs, 2):
                ws_txs.cell(row=row_idx, column=1, value=t.invoice_number).font = font_bold
                ws_txs.cell(row=row_idx, column=2, value=t.timestamp.strftime("%Y-%m-%d %H:%M:%S")).font = font_regular
                ws_txs.cell(row=row_idx, column=3, value=t.payment_method).font = font_regular
                
                c_items = ws_txs.cell(row=row_idx, column=4, value=t.items_count)
                c_sub = ws_txs.cell(row=row_idx, column=5, value=t.subtotal)
                c_gst = ws_txs.cell(row=row_idx, column=6, value=t.gst_amount)
                c_disc = ws_txs.cell(row=row_idx, column=7, value=t.discount_amount)
                c_grand = ws_txs.cell(row=row_idx, column=8, value=t.grand_total)
                c_prof = ws_txs.cell(row=row_idx, column=9, value=t.profit)
                
                c_items.number_format = '#,##0'
                for c in [c_sub, c_gst, c_disc, c_grand, c_prof]:
                    c.number_format = f'"{currency}"#,##0.00'
                    c.font = font_regular
                    c.border = thin_border
                    
                for col_idx in range(1, 10):
                    cell = ws_txs.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    
            # --- Sheet 3: Expenses ---
            ws_exp = wb.create_sheet(title="Expenses Ledger")
            ws_exp.views.sheetView[0].showGridLines = True
            
            exp_headers = ["ID", "Category", "Amount", "Date", "Description"]
            for col_idx, header in enumerate(exp_headers, 1):
                cell = ws_exp.cell(row=1, column=col_idx, value=header)
                cell.font = font_header
                cell.fill = fill_header
                
            expenses = db.query(Expense).filter(Expense.date >= start_date, Expense.date <= end_date).all()
            for row_idx, e in enumerate(expenses, 2):
                ws_exp.cell(row=row_idx, column=1, value=e.id).font = font_regular
                ws_exp.cell(row=row_idx, column=2, value=e.category).font = font_bold
                c_amt = ws_exp.cell(row=row_idx, column=3, value=e.amount)
                c_amt.number_format = f'"{currency}"#,##0.00'
                c_amt.font = font_regular
                ws_exp.cell(row=row_idx, column=4, value=str(e.date)).font = font_regular
                ws_exp.cell(row=row_idx, column=5, value=e.description or "").font = font_regular
                
                for col_idx in range(1, 6):
                    ws_exp.cell(row=row_idx, column=col_idx).border = thin_border
                    
            # Auto-fit column widths for all sheets
            for sheet in wb.worksheets:
                for col in sheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        # Skip title cell in summary sheet when calculating width
                        if sheet.title == "Executive Summary" and cell.coordinate in ["A1", "A2"]:
                            continue
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                    
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            headers = {
                'Content-Disposition': f'attachment; filename="report_{report_type.lower()}_{start_date.strftime("%Y%m%d")}.xlsx"'
            }
            return StreamingResponse(
                output,
                headers=headers,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")
            
    # 3. GENERATE CSV TX LOG
    elif file_format.lower() == "csv":
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            writer.writerow(["Invoice Number", "Timestamp", "Payment Method", "Items Count", "Subtotal", "GST Amount", "Discount", "Grand Total", "Profit"])
            
            txs = db.query(Transaction).filter(
                Transaction.timestamp >= start_date, Transaction.timestamp <= datetime.combine(end_date, datetime.max.time())
            ).all()
            
            for t in txs:
                writer.writerow([
                    t.invoice_number,
                    t.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    t.payment_method,
                    t.items_count,
                    t.subtotal,
                    t.gst_amount,
                    t.discount_amount,
                    t.grand_total,
                    t.profit
                ])
                
            output.seek(0)
            headers = {
                'Content-Disposition': f'attachment; filename="transactions_{start_date.strftime("%Y%m%d")}.csv"'
            }
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8')),
                headers=headers,
                media_type='text/csv'
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"CSV generation failed: {str(e)}")
            
    else:
        raise HTTPException(status_code=400, detail="Invalid format specified. Choose 'pdf', 'excel', or 'csv'")
